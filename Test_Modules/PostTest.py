# import packages
import PySimpleGUI as sg
import datetime
import shutil
from pathlib import Path
import openpyxl
import sys
import os
import dotenv

def run(wo, serial, operator, date, R1, R2, R3, R4, V1, V2, V3, V4, V5, poe1, poe2, poe3, poe4):
    #open excel
    templatePath = 'Z:/05. Manufacturing/60. Uncontrolled/Troubleshoot/Dai/MIS Program/MIS-4251_Func_Test/727-4251 DataSheet.xlsx'
    wb = openpyxl.load_workbook('templatePath')
    ws = wb['Test Report']
    
    #pull info from .env
    dotenv_file = dotenv.find_dotenv()
    dotenv.load_dotenv(dotenv_file)
    
    #write info
    ws['C3'] = serial
    ws['J3'] = operator
    ws['J4'] = date
    #DC PS
    ws['H8'] = os.getenv('DC_MODEL')
    ws['J8'] = os.getenv('DC_EQID')
    ws['L8'] = os.getenv('DC_LAST_CAL')
    ws['M8'] = os.getenv('DC_NEXT_CAL')
    #Multimeter
    ws['H10'] = os.getenv('DMM_MODEL')
    ws['J10'] = os.getenv('DMM_EQID')
    ws['L10'] = os.getenv('DMM_LAST_CAL')
    ws['M10'] = os.getenv('DMM_NEXT_CAL')
    
    # Write the record
    ws['I16'] = R1
    ws['I17'] = R2
    ws['I18'] = R3
    ws['I19'] = R4
    ws['I20'] = V1
    ws['I21'] = V2
    ws['I22'] = V3
    ws['I23'] = V4
    ws['I24'] = V5
    ws['I25'] = poe1
    ws['I26'] = poe2
    ws['I27'] = poe3
    ws['I28'] = poe4
    
    #save
    wb.save(serial + ".xlsx")
    
    #check work order
    if wo.isnumeric():
        Path("Z:/05. Manufacturing/20. Test/400 records/Test Records/727/727-4239/WO#" + wo).mkdir(parents=True, exist_ok=True)
    else:
        Path("Z:/05. Manufacturing/20. Test/400 records/Test Records/727/727-4239/" + wo).mkdir(parents=True, exist_ok=True)
    
    
    src_folder = r"Z:\05. Manufacturing\60. Uncontrolled\Troubleshoot\Phat\MIS\727-4239\\"
    
    #check work order
    if wo.isnumeric():
        dst_folder = r"Z:\05. Manufacturing\20. Test\400 records\Test Records\727\727-4239\WO#" + wo + "//"
    else:
        dst_folder = r"Z:\05. Manufacturing\20. Test\400 records\Test Records\727\727-4239\\" + wo + "//"
        
    file_name = serial + ".xlsx"
    shutil.move(src_folder + file_name, dst_folder + file_name)
    
    return True